from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from difflib import SequenceMatcher
from typing import Any, Iterable
from uuid import uuid4
from xml.etree import ElementTree

from botocore.exceptions import BotoCoreError, ClientError
from flask import current_app
from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage

from meeting_assistant.services.knowledge_service import KnowledgeService
from meeting_assistant.utils.exceptions import DatabaseError, ResourceNotFoundError, ValidationError

_QUESTION_HEADERS = {"question", "questions", "interview question", "prompt"}
_ANSWER_HEADERS = {"answer", "answers", "prepared answer", "response", "suggested answer"}

# Meeting Materials are indexed when they are saved so Live Q&A does not
# download and parse every selected document for every incoming question.
_MATERIAL_INDEX_VERSION = 2
_MAX_MATERIAL_INDEX_ENTRIES = 180
_MAX_MATERIAL_INDEX_BYTES = 180_000
_MAX_MATERIAL_ENTRY_CHARACTERS = 3_000
_MAX_FACT_CHUNK_CHARACTERS = 1_400

_CONTEXT_STOP_WORDS = {
    "a", "about", "an", "and", "are", "as", "at", "be", "been", "being",
    "by", "can", "could", "did", "do", "does", "for", "from", "had", "has",
    "have", "how", "i", "in", "is", "it", "made", "make", "me", "of", "on",
    "or", "our", "please", "regarding", "that", "the", "this", "to", "we",
    "were", "what", "when", "where", "which", "who", "why", "with", "would",
    "you", "your",
}

_CONTEXT_TOKEN_ALIASES = {
    "revenues": "revenue",
    "sales": "sale",
    "sold": "sale",
    "selling": "sale",
    "costs": "cost",
    "expenses": "expense",
    "units": "unit",
    "prices": "price",
    "profits": "profit",
    "margins": "margin",
}

_FINANCIAL_QUERY_TOKENS = {"revenue", "sale", "cost", "expense", "price", "profit", "margin"}
_FINANCIAL_CONTEXT_TOKENS = _FINANCIAL_QUERY_TOKENS | {"unit", "quantity", "amount", "total"}

# Question matching intentionally ignores low-information wording and normalizes
# common interview paraphrases. This lets a prepared question such as
# "What is your expected salary?" match "What salary do you expect?" without
# treating every question containing the word "salary" as equivalent.
_QUESTION_STOP_WORDS = {
    "a", "about", "an", "and", "are", "as", "at", "be", "being", "can", "could",
    "biggest", "describe", "did", "do", "does", "for", "from", "greatest",
    "have", "how", "i", "in",
    "is", "it", "main", "me", "most", "my", "of", "on", "or", "please",
    "say", "strongest", "tell", "that",
    "the", "this", "through", "to", "us", "walk", "was", "were", "what",
    "when", "where",
    "which", "who", "why", "should", "will", "willing", "with", "would",
    "you", "your",
}

_TOKEN_ALIASES = {
    "compensation": "salary",
    "pay": "salary",
    "remuneration": "salary",
    "expected": "expect",
    "expectation": "expect",
    "expectations": "expect",
    "expecting": "expect",
    "desired": "expect",
    "targeted": "target",
    "targeting": "target",
    "seeking": "seek",
    "looking": "look",
    "currently": "current",
    "now": "current",
    "presently": "current",
    "previous": "previous",
    "previously": "previous",
    "prior": "previous",
    "past": "previous",
    "negotiable": "negotiate",
    "negotiation": "negotiate",
    "negotiating": "negotiate",
    "flexibility": "flexible",
    "strengths": "strength",
    "quality": "strength",
    "qualities": "strength",
    "weaknesses": "weakness",
    "positions": "role",
    "position": "role",
    "jobs": "role",
    "job": "role",
    "companies": "company",
    "organizations": "company",
    "organisation": "company",
    "organisations": "company",
    "employer": "company",
    "employers": "company",
    "skills": "skill",
    "qualifications": "qualification",
    "requirements": "requirement",
    "responsibilities": "responsibility",
    "motivated": "motivate",
    "motivation": "motivate",
    "interested": "interest",
    "interests": "interest",
    "yourself": "self_intro",
    "background": "self_intro",
}

# Each family contains mutually different intents. A candidate is rejected when
# the asked question and prepared question clearly select different members of
# the same family. This is especially important for salary questions.
_INTENT_FAMILIES = {
    "salary_stage": {
        "expected": {"expect", "target", "seek", "look", "want"},
        "current": {"current"},
        "previous": {"previous", "former", "last"},
        "negotiation": {"negotiate", "flexible"},
    },
    "personal_trait": {
        "strength": {"strength"},
        "weakness": {"weakness"},
    },
}


class MeetingMaterialsService:
    def __init__(self) -> None:
        self.knowledge = KnowledgeService()
        self.repository = self.knowledge.repository
        self.file_store = self.knowledge.file_store

    def list_meetings(self, user_id: str, *, include_completed: bool = False) -> list[dict[str, Any]]:
        try:
            items = self.repository.list_meetings(user_id)
            active_id = self.repository.get_active_meeting_id(user_id)
        except (BotoCoreError, ClientError, OSError) as exc:
            raise DatabaseError("Upcoming meetings could not be loaded.") from exc
        meetings = [self._serialize_meeting(item, active_id) for item in items]
        if not include_completed:
            meetings = [item for item in meetings if item["status"] not in {"completed", "cancelled"}]
        meetings.sort(key=lambda item: (item.get("scheduled_at") or "9999", item.get("created_at") or ""))
        return meetings

    def create_meeting(self, user_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        title = self.knowledge._required_text(payload.get("title"), "Meeting title", 240)
        meeting_id = str(payload.get("id") or uuid4().hex).strip()[:160]
        if not meeting_id:
            meeting_id = uuid4().hex
        now = _utc_now()
        item = {
            "user_id": user_id,
            "item_id": f"meeting#{meeting_id}",
            "entity_type": "meeting_package",
            "meeting_id": meeting_id,
            "title": title,
            "scheduled_at": str(payload.get("scheduled_at") or "").strip()[:80],
            "participants": _string_list(payload.get("participants"), 100, 200),
            "purpose": str(payload.get("purpose") or "").strip()[:2000],
            "status": str(payload.get("status") or ("upcoming" if payload.get("scheduled_at") else "draft")),
            "library_file_ids": [],
            "temporary_files": [],
            "meeting_context": {},
            "material_index": [],
            "material_index_version": _MATERIAL_INDEX_VERSION,
            "material_indexed_at": now,
            "created_at": now,
            "updated_at": now,
            "completed_at": "",
            "completed_meeting_id": "",
        }
        self._save(item)
        if bool(payload.get("activate", True)):
            self.set_active_meeting(user_id, meeting_id)
        return self._serialize_meeting(item, meeting_id if payload.get("activate", True) else self.get_active_meeting_id(user_id))

    def update_meeting(self, user_id: str, meeting_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        item = self._get(user_id, meeting_id)
        for key, maximum in (("title", 240), ("scheduled_at", 80), ("purpose", 2000), ("status", 30), ("completed_at", 80), ("completed_meeting_id", 160)):
            if key in payload:
                value = str(payload.get(key) or "").strip()[:maximum]
                if key == "title" and not value:
                    raise ValidationError("Meeting title is required.")
                item[key] = value
        if "participants" in payload:
            item["participants"] = _string_list(payload.get("participants"), 100, 200)
        item["updated_at"] = _utc_now()
        self._save(item)
        if payload.get("activate") is True:
            self.set_active_meeting(user_id, meeting_id)
        return self._serialize_meeting(item, self.get_active_meeting_id(user_id))

    def delete_meeting(self, user_id: str, meeting_id: str) -> dict[str, Any]:
        item = self._get(user_id, meeting_id)
        normalized_id = str(item.get("meeting_id") or meeting_id)

        # Clear the active package first so Live Q&A never keeps pointing to a
        # meeting that is about to be removed.
        if self.get_active_meeting_id(user_id) == normalized_id:
            self.set_active_meeting(user_id, "")

        try:
            deleted = self.repository.delete_meeting(user_id, normalized_id)
        except ResourceNotFoundError:
            raise
        except (BotoCoreError, ClientError, OSError) as exc:
            raise DatabaseError("The meeting package could not be deleted.") from exc

        # Temporary meeting files belong exclusively to this package. Metadata
        # deletion is the primary operation; storage cleanup is best-effort so
        # an unavailable object store does not make the deleted meeting reappear.
        for record in deleted.get("temporary_files") or []:
            object_key = str(record.get("object_key") or "").strip()
            if not object_key:
                continue
            try:
                self.file_store.delete(object_key)
            except Exception:
                current_app.logger.exception(
                    "Could not delete temporary file for removed meeting package %s",
                    normalized_id,
                )

        return self._serialize_meeting(deleted)

    def get_materials(self, user_id: str, meeting_id: str) -> dict[str, Any]:
        item = self._get(user_id, meeting_id)
        return self._serialize_materials(item)

    def save_materials(self, user_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        meeting_id = str(payload.get("meeting_id") or "").strip()
        item = self._get(user_id, meeting_id)
        requested_ids = list(dict.fromkeys(_string_list(payload.get("library_file_ids") or payload.get("file_ids"), 500, 160)))
        valid_ids = []
        for file_id in requested_ids:
            if self.repository.get_file(user_id, file_id):
                valid_ids.append(file_id)
            else:
                raise ValidationError("One of the selected documents no longer exists.")
        item["library_file_ids"] = valid_ids
        if "meeting_context" in payload and isinstance(payload.get("meeting_context"), dict):
            item["meeting_context"] = _clean_context(payload["meeting_context"])
        self._rebuild_material_index(user_id, item)
        item["updated_at"] = _utc_now()
        self._save(item)
        if payload.get("activate", True):
            self.set_active_meeting(user_id, meeting_id)
        return self._serialize_materials(item)

    def save_meeting_context(self, user_id: str, meeting_id: str, context: dict[str, Any]) -> dict[str, Any]:
        item = self._get(user_id, meeting_id)
        item["meeting_context"] = _clean_context(context)
        item["updated_at"] = _utc_now()
        self._save(item)
        return dict(item["meeting_context"])

    def upload_temporary_files(self, user_id: str, meeting_id: str, uploads: Iterable[FileStorage]) -> list[dict[str, Any]]:
        item = self._get(user_id, meeting_id)
        selected = [upload for upload in uploads if upload and upload.filename]
        if not selected:
            raise ValidationError("Choose at least one temporary file.")
        temporary = list(item.get("temporary_files") or [])
        added: list[dict[str, Any]] = []
        try:
            for upload in selected:
                original_name, stored_name, extension, content_type, content = self.knowledge._prepare_upload(upload)
                file_id = uuid4().hex
                object_key = self._temporary_object_key(user_id, meeting_id, file_id, stored_name)
                self.file_store.put(object_key, content, content_type)
                record = {
                    "file_id": file_id,
                    "id": file_id,
                    "name": original_name,
                    "filename": original_name,
                    "extension": extension,
                    "content_type": content_type,
                    "size_bytes": len(content),
                    "object_key": object_key,
                    "added_at": _utc_now(),
                }
                temporary.append(record)
                added.append(record)
            item["temporary_files"] = temporary
            self._rebuild_material_index(user_id, item)
            item["updated_at"] = _utc_now()
            self._save(item)
        except Exception:
            for record in added:
                try:
                    self.file_store.delete(record["object_key"])
                except Exception:
                    current_app.logger.exception("Could not roll back temporary meeting file")
            raise
        return [self._serialize_temporary_file(record) for record in added]

    def delete_temporary_file(self, user_id: str, meeting_id: str, file_id: str) -> None:
        item = self._get(user_id, meeting_id)
        files = list(item.get("temporary_files") or [])
        target = next((record for record in files if str(record.get("file_id") or record.get("id")) == file_id), None)
        if not target:
            raise ResourceNotFoundError("Temporary file not found.")
        item["temporary_files"] = [record for record in files if record is not target]
        self._rebuild_material_index(user_id, item)
        item["updated_at"] = _utc_now()
        self._save(item)
        self.file_store.delete(str(target.get("object_key") or ""))

    def clear_temporary_files(self, user_id: str, meeting_id: str) -> None:
        item = self._get(user_id, meeting_id)
        files = list(item.get("temporary_files") or [])
        item["temporary_files"] = []
        self._rebuild_material_index(user_id, item)
        item["updated_at"] = _utc_now()
        self._save(item)
        for record in files:
            try:
                self.file_store.delete(str(record.get("object_key") or ""))
            except Exception:
                current_app.logger.exception("Could not delete temporary meeting file")

    def get_active_meeting_id(self, user_id: str) -> str:
        try:
            return self.repository.get_active_meeting_id(user_id)
        except (BotoCoreError, ClientError, OSError) as exc:
            raise DatabaseError("The active meeting package could not be loaded.") from exc

    def set_active_meeting(self, user_id: str, meeting_id: str) -> str:
        normalized = str(meeting_id or "").strip()
        if normalized:
            item = self._get(user_id, normalized)
            if str(item.get("status") or "") in {"completed", "cancelled"}:
                raise ValidationError("A completed or cancelled meeting cannot be active.")
        try:
            self.repository.set_active_meeting_id(user_id, normalized)
        except (BotoCoreError, ClientError, OSError) as exc:
            raise DatabaseError("The active meeting package could not be saved.") from exc
        return normalized

    def complete_meeting(self, user_id: str, prepared_meeting_id: str, completed_meeting_id: str) -> None:
        if not prepared_meeting_id:
            return
        item = self._get(user_id, prepared_meeting_id)
        now = _utc_now()
        item.update({"status": "completed", "completed_at": now, "completed_meeting_id": completed_meeting_id, "updated_at": now})
        self._save(item)
        if self.get_active_meeting_id(user_id) == prepared_meeting_id:
            self.set_active_meeting(user_id, "")

    def find_prepared_answer(self, user_id: str, question: str, meeting_id: str = "") -> dict[str, Any] | None:
        selected_id, meeting = self._resolve_meeting(user_id, meeting_id)
        if not meeting:
            return None
        normalized_question = _normalize_question(question)
        if not normalized_question:
            return None

        candidates = [
            entry
            for entry in self._ensure_material_index(user_id, meeting)
            if str(entry.get("source_type") or "") == "prepared_answer"
        ]
        ranked: list[tuple[float, str, dict[str, Any]]] = []
        for candidate in candidates:
            score, confidence = _question_match(
                normalized_question,
                str(candidate.get("question") or ""),
            )
            if confidence != "none":
                ranked.append((score, confidence, candidate))

        if not ranked:
            return None

        ranked.sort(key=lambda item: item[0], reverse=True)
        score, confidence, candidate = ranked[0]

        # Avoid returning a prepared answer verbatim when two different prepared
        # questions are almost equally plausible. The best match remains useful
        # as authoritative context, but the AI may adapt it to the exact wording.
        if confidence == "high" and len(ranked) > 1:
            second_score, _, second_candidate = ranked[1]
            different_question = (
                _normalize_question(candidate.get("question", ""))
                != _normalize_question(second_candidate.get("question", ""))
            )
            if different_question and score - second_score < 0.04:
                confidence = "medium"

        return {
            **candidate,
            "meeting_id": selected_id,
            "meeting_title": str(meeting.get("title") or ""),
            "match_score": round(score, 4),
            "match_confidence": confidence,
            "use_verbatim": confidence == "high",
        }

    def find_relevant_context(
        self,
        user_id: str,
        question: str,
        meeting_id: str = "",
        *,
        limit: int = 4,
    ) -> list[dict[str, Any]]:
        """Returns factual excerpts from the active Meeting Materials.

        This is intentionally separate from prepared-answer matching. It lets
        ordinary statements, spreadsheet rows, and paragraphs support Live Q&A
        calculations and factual answers even when the document was not written
        as a question-and-answer list.
        """
        selected_id, meeting = self._resolve_meeting(user_id, meeting_id)
        if not meeting or not str(question or "").strip():
            return []

        ranked: list[tuple[float, dict[str, Any]]] = []
        for entry in self._ensure_material_index(user_id, meeting):
            source_type = str(entry.get("source_type") or "")
            if source_type not in {"document_context", "prepared_answer"}:
                continue
            score = _context_match_score(question, entry)
            if score <= 0:
                continue
            ranked.append((score, entry))

        ranked.sort(key=lambda item: item[0], reverse=True)
        results: list[dict[str, Any]] = []
        seen_text: set[str] = set()
        for score, entry in ranked:
            context_text = _entry_context_text(entry)
            normalized_text = _normalize_question(context_text)
            if not context_text or normalized_text in seen_text:
                continue
            seen_text.add(normalized_text)
            results.append({
                **entry,
                "text": context_text,
                "meeting_id": selected_id,
                "meeting_title": str(meeting.get("title") or ""),
                "match_score": round(score, 4),
            })
            if len(results) >= max(1, min(int(limit), 8)):
                break
        return results

    def _resolve_meeting(
        self,
        user_id: str,
        meeting_id: str = "",
    ) -> tuple[str, dict[str, Any] | None]:
        selected_id = str(meeting_id or "").strip() or self.get_active_meeting_id(user_id)
        if not selected_id:
            return "", None
        try:
            return selected_id, self._get(user_id, selected_id)
        except ResourceNotFoundError:
            return selected_id, None

    def _ensure_material_index(
        self,
        user_id: str,
        meeting: dict[str, Any],
    ) -> list[dict[str, Any]]:
        if (
            int(meeting.get("material_index_version") or 0) == _MATERIAL_INDEX_VERSION
            and isinstance(meeting.get("material_index"), list)
        ):
            return list(meeting.get("material_index") or [])

        # Existing meeting packages are migrated lazily on their first Live Q&A
        # request. Subsequent requests use the stored compact index and do not
        # reopen the selected files.
        self._rebuild_material_index(user_id, meeting)
        meeting["updated_at"] = _utc_now()
        self._save(meeting)
        return list(meeting.get("material_index") or [])

    def _rebuild_material_index(
        self,
        user_id: str,
        meeting: dict[str, Any],
    ) -> None:
        entries: list[dict[str, Any]] = []
        total_bytes = 0

        def add_document(file_item: dict[str, Any], content: bytes) -> None:
            nonlocal total_bytes
            try:
                extracted = _extract_material_entries(file_item, content)
            except Exception:
                current_app.logger.exception(
                    "Could not index Meeting Material %s",
                    file_item.get("file_id") or file_item.get("filename") or file_item.get("name"),
                )
                return

            for raw_entry in extracted:
                entry = _sanitize_index_entry(raw_entry)
                entry_size = _index_entry_byte_count(entry)
                if not entry or entry_size <= 0:
                    continue
                if len(entries) >= _MAX_MATERIAL_INDEX_ENTRIES:
                    return
                if total_bytes + entry_size > _MAX_MATERIAL_INDEX_BYTES:
                    return
                entries.append(entry)
                total_bytes += entry_size

        for file_id in meeting.get("library_file_ids") or []:
            file_item = self.repository.get_file(user_id, str(file_id))
            if not file_item:
                continue
            try:
                content = self.file_store.get(str(file_item.get("object_key") or ""))
            except Exception:
                current_app.logger.exception("Could not load Meeting Material %s", file_id)
                continue
            add_document(file_item, content)

        for file_item in meeting.get("temporary_files") or []:
            try:
                content = self.file_store.get(str(file_item.get("object_key") or ""))
            except Exception:
                current_app.logger.exception("Could not load temporary Meeting Material")
                continue
            add_document(file_item, content)

        meeting["material_index"] = entries
        meeting["material_index_version"] = _MATERIAL_INDEX_VERSION
        meeting["material_indexed_at"] = _utc_now()

    def _get(self, user_id: str, meeting_id: str) -> dict[str, Any]:
        normalized = str(meeting_id or "").strip()
        if not normalized:
            raise ValidationError("meeting_id is required.")
        try:
            item = self.repository.get_meeting(user_id, normalized)
        except (BotoCoreError, ClientError, OSError) as exc:
            raise DatabaseError("The meeting package could not be loaded.") from exc
        if not item:
            raise ResourceNotFoundError("Meeting package not found.")
        return item

    def _save(self, item: dict[str, Any]) -> None:
        try:
            self.repository.upsert_meeting(item)
        except (BotoCoreError, ClientError, OSError) as exc:
            raise DatabaseError("The meeting package could not be saved.") from exc

    @staticmethod
    def _serialize_meeting(item: dict[str, Any], active_id: str = "") -> dict[str, Any]:
        meeting_id = str(item.get("meeting_id") or "")
        return {
            "id": meeting_id,
            "meeting_id": meeting_id,
            "title": str(item.get("title") or "Untitled meeting"),
            "scheduled_at": str(item.get("scheduled_at") or ""),
            "participants": list(item.get("participants") or []),
            "purpose": str(item.get("purpose") or ""),
            "status": str(item.get("status") or "draft"),
            "created_at": str(item.get("created_at") or ""),
            "updated_at": str(item.get("updated_at") or ""),
            "completed_at": str(item.get("completed_at") or ""),
            "completed_meeting_id": str(item.get("completed_meeting_id") or ""),
            "active": meeting_id == active_id,
            "library_file_count": len(item.get("library_file_ids") or []),
            "temporary_file_count": len(item.get("temporary_files") or []),
        }

    def _serialize_materials(self, item: dict[str, Any]) -> dict[str, Any]:
        return {
            "meeting_id": str(item.get("meeting_id") or ""),
            "library_file_ids": list(item.get("library_file_ids") or []),
            "temporary_files": [self._serialize_temporary_file(record) for record in item.get("temporary_files") or []],
            "meeting_context": _clean_context(item.get("meeting_context") or {}),
        }

    @staticmethod
    def _serialize_temporary_file(record: dict[str, Any]) -> dict[str, Any]:
        file_id = str(record.get("file_id") or record.get("id") or "")
        return {"id": file_id, "file_id": file_id, "name": str(record.get("name") or record.get("filename") or "Temporary file"), "filename": str(record.get("filename") or record.get("name") or "Temporary file"), "extension": str(record.get("extension") or ""), "content_type": str(record.get("content_type") or ""), "size_bytes": int(record.get("size_bytes") or 0), "added_at": str(record.get("added_at") or "")}

    @staticmethod
    def _temporary_object_key(user_id: str, meeting_id: str, file_id: str, stored_name: str) -> str:
        import hashlib
        owner_hash = hashlib.sha256(user_id.encode("utf-8")).hexdigest()[:24]
        return f"knowledge/{owner_hash}/meetings/{meeting_id}/temporary/{file_id}/{stored_name}"


def _extract_material_entries(item: dict[str, Any], content: bytes) -> list[dict[str, Any]]:
    extension = str(item.get("extension") or "").lower()
    filename = str(item.get("display_name") or item.get("filename") or item.get("name") or "Document")
    prepared = _extract_candidates(item, content)
    if extension == "xlsx":
        facts = _xlsx_fact_chunks(filename, content)
    elif extension == "xls":
        facts = _xls_fact_chunks(filename, content)
    elif extension == "pdf":
        facts = _pdf_fact_chunks(filename, content)
    elif extension in {"txt", "md"}:
        facts = _text_fact_chunks(filename, content.decode("utf-8-sig", errors="replace"))
    elif extension == "docx":
        facts = _docx_fact_chunks(filename, content)
    else:
        facts = []
    return prepared + facts


def _extract_candidates(item: dict[str, Any], content: bytes) -> list[dict[str, Any]]:
    extension = str(item.get("extension") or "").lower()
    filename = str(item.get("display_name") or item.get("filename") or item.get("name") or "Document")
    if extension == "xlsx":
        return _xlsx_candidates(filename, content)
    if extension == "xls":
        return _xls_candidates(filename, content)
    if extension == "pdf":
        return _text_candidates(filename, _pdf_text(content))
    if extension in {"txt", "md"}:
        return _text_candidates(filename, content.decode("utf-8-sig", errors="replace"))
    if extension == "docx":
        return _text_candidates(filename, _docx_text(content))
    return []


def _xlsx_candidates(filename: str, content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    results: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            rows = [[str(value).strip() if value is not None else "" for value in row] for row in sheet.iter_rows(values_only=True)]
            question_index = answer_index = None
            start = 0
            for index, row in enumerate(rows[:20]):
                lowered = [value.casefold() for value in row]
                q = next((i for i, value in enumerate(lowered) if value in _QUESTION_HEADERS), None)
                a = next((i for i, value in enumerate(lowered) if value in _ANSWER_HEADERS), None)
                if q is not None and a is not None:
                    question_index, answer_index, start = q, a, index + 1
                    break
            for row_number, row in enumerate(rows[start:], start=start + 1):
                if question_index is not None and answer_index is not None:
                    question = row[question_index] if question_index < len(row) else ""
                    answer = row[answer_index] if answer_index < len(row) else ""
                else:
                    values = [value for value in row if value]
                    if len(values) < 2:
                        continue
                    question, answer = values[0], values[1]
                    if "?" not in question and not re.match(r"^(tell|describe|explain|why|how|what|when|where|who|give|walk)\b", question, re.I):
                        continue
                if question and answer:
                    results.append({"question": question, "answer": answer, "source_name": filename, "source_detail": f"{sheet.title} · row {row_number}", "source_type": "prepared_answer"})
    finally:
        workbook.close()
    return results



def _xls_candidates(filename: str, content: bytes) -> list[dict[str, Any]]:
    import xlrd
    workbook = xlrd.open_workbook(file_contents=content)
    results: list[dict[str, Any]] = []
    for sheet in workbook.sheets():
        rows = [[str(sheet.cell_value(row, column)).strip() for column in range(sheet.ncols)] for row in range(sheet.nrows)]
        question_index = answer_index = None
        start = 0
        for index, row in enumerate(rows[:20]):
            lowered = [value.casefold() for value in row]
            q = next((i for i, value in enumerate(lowered) if value in _QUESTION_HEADERS), None)
            a = next((i for i, value in enumerate(lowered) if value in _ANSWER_HEADERS), None)
            if q is not None and a is not None:
                question_index, answer_index, start = q, a, index + 1
                break
        for row_number, row in enumerate(rows[start:], start=start + 1):
            if question_index is not None and answer_index is not None:
                question = row[question_index] if question_index < len(row) else ""
                answer = row[answer_index] if answer_index < len(row) else ""
            else:
                values = [value for value in row if value]
                if len(values) < 2:
                    continue
                question, answer = values[0], values[1]
                if "?" not in question and not re.match(r"^(tell|describe|explain|why|how|what|when|where|who|give|walk)\b", question, re.I):
                    continue
            if question and answer:
                results.append({"question": question, "answer": answer, "source_name": filename, "source_detail": f"{sheet.name} · row {row_number}", "source_type": "prepared_answer"})
    return results


def _pdf_text(content: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _text_candidates(filename: str, text: str) -> list[dict[str, Any]]:
    results = []
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for index, line in enumerate(lines):
        match = re.match(r"^(?:q(?:uestion)?\s*[:\-]\s*)?(.+?\?)\s*(?:a(?:nswer)?\s*[:\-]\s*)?(.+)$", line, re.I)
        if match:
            results.append({"question": match.group(1).strip(), "answer": match.group(2).strip(), "source_name": filename, "source_detail": f"line {index + 1}", "source_type": "prepared_answer"})
        elif line.endswith("?") and index + 1 < len(lines):
            results.append({"question": line, "answer": lines[index + 1], "source_name": filename, "source_detail": f"lines {index + 1}-{index + 2}", "source_type": "prepared_answer"})
    return results


def _docx_text(content: bytes) -> str:
    import zipfile
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    return "\n".join(node.text or "" for node in root.iter() if node.tag.endswith("}t"))



def _xlsx_fact_chunks(filename: str, content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    results: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            headers: list[str] = []
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [_cell_text(value) for value in row]
                if not any(values):
                    continue
                if not headers and _looks_like_header_row(values):
                    headers = values
                text = _spreadsheet_row_text(values, headers if row_number > 1 else [])
                if text:
                    results.append(_fact_entry(filename, f"{sheet.title} · row {row_number}", text))
    finally:
        workbook.close()
    return results


def _xls_fact_chunks(filename: str, content: bytes) -> list[dict[str, Any]]:
    import xlrd
    workbook = xlrd.open_workbook(file_contents=content)
    results: list[dict[str, Any]] = []
    for sheet in workbook.sheets():
        headers: list[str] = []
        for row_index in range(sheet.nrows):
            values = [_cell_text(sheet.cell_value(row_index, column)) for column in range(sheet.ncols)]
            if not any(values):
                continue
            if not headers and _looks_like_header_row(values):
                headers = values
            text = _spreadsheet_row_text(values, headers if row_index > 0 else [])
            if text:
                results.append(_fact_entry(filename, f"{sheet.name} · row {row_index + 1}", text))
    return results


def _pdf_fact_chunks(filename: str, content: bytes) -> list[dict[str, Any]]:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(content))
    results: list[dict[str, Any]] = []
    for page_number, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""
        for chunk_number, chunk in enumerate(_chunk_text(page_text), start=1):
            detail = f"page {page_number}"
            if chunk_number > 1:
                detail += f" · part {chunk_number}"
            results.append(_fact_entry(filename, detail, chunk))
    return results


def _text_fact_chunks(filename: str, text: str) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    lines = text.splitlines()
    paragraph_lines: list[str] = []
    paragraph_start = 1

    def flush(end_line: int) -> None:
        nonlocal paragraph_lines, paragraph_start
        paragraph = " ".join(line.strip() for line in paragraph_lines if line.strip()).strip()
        if paragraph:
            for chunk_number, chunk in enumerate(_chunk_text(paragraph), start=1):
                detail = f"lines {paragraph_start}-{end_line}" if end_line != paragraph_start else f"line {paragraph_start}"
                if chunk_number > 1:
                    detail += f" · part {chunk_number}"
                results.append(_fact_entry(filename, detail, chunk))
        paragraph_lines = []

    for line_number, line in enumerate(lines, start=1):
        if line.strip():
            if not paragraph_lines:
                paragraph_start = line_number
            paragraph_lines.append(line)
        else:
            flush(line_number - 1)
    flush(len(lines))
    return results


def _docx_fact_chunks(filename: str, content: bytes) -> list[dict[str, Any]]:
    import zipfile
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    paragraphs: list[str] = []
    for paragraph in root.iter():
        if not paragraph.tag.endswith("}p"):
            continue
        text = "".join(node.text or "" for node in paragraph.iter() if node.tag.endswith("}t")).strip()
        if text:
            paragraphs.append(text)
    results: list[dict[str, Any]] = []
    for paragraph_number, paragraph in enumerate(paragraphs, start=1):
        for chunk_number, chunk in enumerate(_chunk_text(paragraph), start=1):
            detail = f"paragraph {paragraph_number}"
            if chunk_number > 1:
                detail += f" · part {chunk_number}"
            results.append(_fact_entry(filename, detail, chunk))
    return results


def _fact_entry(filename: str, detail: str, text: str) -> dict[str, Any]:
    return {
        "text": str(text or "").strip()[:_MAX_MATERIAL_ENTRY_CHARACTERS],
        "source_name": filename,
        "source_detail": detail,
        "source_type": "document_context",
    }


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_like_header_row(values: list[str]) -> bool:
    nonempty = [value for value in values if value]
    if not nonempty:
        return False
    alpha_cells = sum(bool(re.search(r"[A-Za-z]", value)) and not re.search(r"\d", value) for value in nonempty)
    return alpha_cells >= max(1, len(nonempty) // 2)


def _spreadsheet_row_text(values: list[str], headers: list[str]) -> str:
    pairs: list[str] = []
    for index, value in enumerate(values):
        if not value:
            continue
        header = headers[index].strip() if index < len(headers) else ""
        if header and header.casefold() != value.casefold():
            pairs.append(f"{header}: {value}")
        else:
            pairs.append(value)
    return "; ".join(pairs)[:_MAX_FACT_CHUNK_CHARACTERS]


def _chunk_text(text: str) -> list[str]:
    normalized = re.sub(r"\s+", " ", str(text or "")).strip()
    if not normalized:
        return []
    sentences = re.split(r"(?<=[.!?])\s+", normalized)
    chunks: list[str] = []
    current = ""
    for sentence in sentences:
        sentence = sentence.strip()
        if not sentence:
            continue
        if len(sentence) > _MAX_FACT_CHUNK_CHARACTERS:
            if current:
                chunks.append(current)
                current = ""
            for offset in range(0, len(sentence), _MAX_FACT_CHUNK_CHARACTERS):
                chunks.append(sentence[offset: offset + _MAX_FACT_CHUNK_CHARACTERS])
            continue
        candidate = f"{current} {sentence}".strip()
        if current and len(candidate) > _MAX_FACT_CHUNK_CHARACTERS:
            chunks.append(current)
            current = sentence
        else:
            current = candidate
    if current:
        chunks.append(current)
    return chunks


def _sanitize_index_entry(entry: dict[str, Any]) -> dict[str, Any]:
    source_type = str(entry.get("source_type") or "").strip()
    if source_type == "prepared_answer":
        question = str(entry.get("question") or "").strip()[:1_000]
        answer = str(entry.get("answer") or "").strip()[:_MAX_MATERIAL_ENTRY_CHARACTERS]
        if not question or not answer:
            return {}
        sanitized = {"question": question, "answer": answer, "source_type": source_type}
    elif source_type == "document_context":
        content = str(entry.get("text") or "").strip()[:_MAX_MATERIAL_ENTRY_CHARACTERS]
        if not content:
            return {}
        sanitized = {"text": content, "source_type": source_type}
    else:
        return {}
    sanitized["source_name"] = str(entry.get("source_name") or "Meeting material").strip()[:500]
    sanitized["source_detail"] = str(entry.get("source_detail") or "").strip()[:500]
    return sanitized


def _index_entry_byte_count(entry: dict[str, Any]) -> int:
    return sum(len(str(value).encode("utf-8")) for value in entry.values())


def _entry_context_text(entry: dict[str, Any]) -> str:
    if str(entry.get("source_type") or "") == "prepared_answer":
        question = str(entry.get("question") or "").strip()
        answer = str(entry.get("answer") or "").strip()
        return f"Question: {question}\nPrepared answer: {answer}".strip()
    return str(entry.get("text") or "").strip()


def _context_tokens(value: str) -> set[str]:
    tokens: set[str] = set()
    for token in re.findall(r"[A-Za-z0-9][A-Za-z0-9_.-]*", str(value or "").casefold()):
        if token in _CONTEXT_STOP_WORDS:
            continue
        canonical = _CONTEXT_TOKEN_ALIASES.get(token, token)
        if len(canonical) > 3 and canonical.endswith("s") and not canonical.endswith("ss"):
            canonical = canonical[:-1]
        tokens.add(_CONTEXT_TOKEN_ALIASES.get(canonical, canonical))
    return tokens


def _identifier_tokens(value: str) -> set[str]:
    return {
        token.casefold()
        for token in re.findall(r"\b(?=[A-Za-z0-9_-]*[A-Za-z])(?=[A-Za-z0-9_-]*\d)[A-Za-z0-9][A-Za-z0-9_-]{2,}\b", str(value or ""))
    }


def _context_match_score(question: str, entry: dict[str, Any]) -> float:
    context_text = _entry_context_text(entry)
    question_tokens = _context_tokens(question)
    context_tokens = _context_tokens(context_text)
    if not question_tokens or not context_tokens:
        return 0.0

    question_ids = _identifier_tokens(question)
    context_ids = _identifier_tokens(context_text)
    shared_ids = question_ids & context_ids
    if question_ids and not shared_ids:
        return 0.0

    overlap = question_tokens & context_tokens
    query_coverage = len(overlap) / len(question_tokens)
    context_precision = len(overlap) / min(len(context_tokens), 20)
    score = (0.56 * query_coverage) + (0.20 * context_precision)

    if shared_ids:
        # Exact product/project/account identifiers are stronger evidence than
        # broad vocabulary overlap. This is the key signal for BRAC204-like IDs.
        score += 0.55 + min(0.10, 0.05 * len(shared_ids))

    financial_query = bool(question_tokens & _FINANCIAL_QUERY_TOKENS)
    financial_context = bool(context_tokens & _FINANCIAL_CONTEXT_TOKENS)
    has_numeric_evidence = bool(re.search(r"(?:[$€£]\s*\d|\b\d+(?:[.,]\d+)?\b)", context_text))
    asks_amount = bool(re.search(r"\b(how much|total|amount|revenue|cost|price|profit|margin)\b", question, re.I))
    if financial_query and financial_context:
        score += 0.12
    if asks_amount and has_numeric_evidence:
        score += 0.12

    # Without a specific identifier, require meaningful lexical support to avoid
    # attaching generic but unrelated document passages.
    if not shared_ids and (query_coverage < 0.28 or len(overlap) < 2):
        return 0.0
    return min(score, 1.0)

def _normalize_question(value: str) -> str:
    text = re.sub(r"[^\w\s]", " ", str(value or "").casefold())
    return re.sub(r"\s+", " ", text).strip()


def _canonical_token(token: str) -> str:
    token = _TOKEN_ALIASES.get(token, token)
    if token in _TOKEN_ALIASES:
        return _TOKEN_ALIASES[token]

    # Conservative stemming for ordinary plurals and verb forms. Aliases are
    # checked again after stemming so words such as "expectations" normalize
    # predictably without requiring a heavyweight NLP dependency.
    stemmed = token
    if len(stemmed) > 5 and stemmed.endswith("ies"):
        stemmed = stemmed[:-3] + "y"
    elif len(stemmed) > 5 and stemmed.endswith("ing"):
        stemmed = stemmed[:-3]
    elif len(stemmed) > 4 and stemmed.endswith("ed"):
        stemmed = stemmed[:-2]
    elif len(stemmed) > 4 and stemmed.endswith("s") and not stemmed.endswith("ss"):
        stemmed = stemmed[:-1]
    return _TOKEN_ALIASES.get(stemmed, stemmed)


def _meaningful_question_tokens(value: str) -> list[str]:
    normalized = _normalize_question(value)
    tokens = [
        _canonical_token(token)
        for token in normalized.split()
        if token and token not in _QUESTION_STOP_WORDS
    ]

    # Salary expectation phrases have several natural forms. Canonicalize their
    # intent only when salary is actually present, preventing a phrase such as
    # "Why do you want this role?" from being interpreted as salary-related.
    if "salary" in tokens:
        salary_expectation_words = {"target", "seek", "look", "want"}
        if any(token in salary_expectation_words for token in tokens):
            tokens = [
                "expect" if token in salary_expectation_words else token
                for token in tokens
            ]
        if "flexible" in tokens:
            tokens = ["negotiate" if token == "flexible" else token for token in tokens]
        if not any(token in {"current", "previous", "negotiate"} for token in tokens):
            if any(token in {"range", "requirement"} for token in tokens):
                tokens.append("expect")

    if "role" in tokens and any(token in {"want", "motivate", "interest"} for token in tokens):
        tokens = [
            "interest" if token in {"want", "motivate", "interest"} else token
            for token in tokens
        ]

    # Preserve order for the secondary sequence signal while removing duplicates.
    return list(dict.fromkeys(tokens))


def _question_intents(tokens: set[str]) -> dict[str, set[str]]:
    intents: dict[str, set[str]] = {}
    for family, labels in _INTENT_FAMILIES.items():
        if family == "salary_stage" and "salary" not in tokens:
            continue
        matched = {
            label
            for label, vocabulary in labels.items()
            if tokens.intersection(vocabulary)
        }
        if matched:
            intents[family] = matched
    return intents


def _intent_relationship(question_tokens: set[str], candidate_tokens: set[str]) -> tuple[bool, bool]:
    question_intents = _question_intents(question_tokens)
    candidate_intents = _question_intents(candidate_tokens)
    partial = False
    for family in question_intents.keys() & candidate_intents.keys():
        question_labels = question_intents[family]
        candidate_labels = candidate_intents[family]
        if question_labels.isdisjoint(candidate_labels):
            return True, False
        if question_labels != candidate_labels:
            partial = True
    return False, partial


def _question_match(question: str, candidate_question: str) -> tuple[float, str]:
    normalized_question = _normalize_question(question)
    normalized_candidate = _normalize_question(candidate_question)
    if not normalized_question or not normalized_candidate:
        return 0.0, "none"
    if normalized_question == normalized_candidate:
        return 1.0, "high"

    question_tokens = _meaningful_question_tokens(normalized_question)
    candidate_tokens = _meaningful_question_tokens(normalized_candidate)
    question_set = set(question_tokens)
    candidate_set = set(candidate_tokens)

    if not question_set or not candidate_set:
        score = SequenceMatcher(None, normalized_question, normalized_candidate).ratio()
        return (score, "medium") if score >= 0.78 else (score, "none")

    intent_conflict, partial_intent = _intent_relationship(question_set, candidate_set)
    if intent_conflict:
        return 0.0, "none"

    intersection = question_set & candidate_set
    if not intersection:
        return 0.0, "none"

    containment = len(intersection) / min(len(question_set), len(candidate_set))
    jaccard = len(intersection) / len(question_set | candidate_set)
    character_score = SequenceMatcher(None, normalized_question, normalized_candidate).ratio()
    token_order_score = SequenceMatcher(None, question_tokens, candidate_tokens).ratio()
    score = (0.50 * containment) + (0.25 * jaccard) + (0.15 * character_score) + (0.10 * token_order_score)

    # Identical semantic token sets are high confidence even when the words are
    # reordered. This is the important case for "expected salary" versus
    # "salary do you expect". Two or more strongly overlapping core tokens also
    # qualify when no conflicting intent was detected.
    if not partial_intent and question_set == candidate_set:
        return max(score, 0.90), "high"
    if not partial_intent and len(intersection) >= 2 and containment >= 0.80 and jaccard >= 0.55:
        return score, "high"
    if not partial_intent and score >= 0.84 and containment >= 0.70:
        return score, "high"

    # Medium matches remain useful, but should be passed to the model as an
    # authoritative prepared answer rather than returned blindly.
    if score >= 0.58 and containment >= 0.50:
        return score, "medium"
    return score, "none"


def _string_list(value: Any, maximum_items: int, maximum_length: int) -> list[str]:
    values = value if isinstance(value, list) else str(value or "").split(",")
    return [str(item).strip()[:maximum_length] for item in values if str(item).strip()][:maximum_items]


def _clean_context(value: dict[str, Any]) -> dict[str, Any]:
    raw = value if isinstance(value, dict) else {}
    special_instructions = str(
        raw.get("special_instructions")
        or raw.get("meeting_instructions")
        or raw.get("instructions")
        or ""
    ).strip()

    # Preserve meeting contexts saved by older clients that used three
    # separate narrative fields. They are combined into the new single field
    # the first time the context is read or saved.
    if not special_instructions:
        legacy_parts: list[str] = []
        topics = str(raw.get("topics") or "").strip()
        constraints = str(raw.get("constraints") or "").strip()
        free_text = str(raw.get("free_text") or "").strip()
        if topics:
            legacy_parts.append(f"Topics to prioritize: {topics}")
        if constraints:
            legacy_parts.append(f"Constraints or sensitivities: {constraints}")
        if free_text:
            legacy_parts.append(free_text)
        special_instructions = "\n".join(legacy_parts)

    return {
        "objective": str(raw.get("objective") or "").strip()[:4000],
        "participants": str(raw.get("participants") or "").strip()[:4000],
        "special_instructions": special_instructions[:4000],
    }


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()
