from __future__ import annotations

import io
import json
import re
from datetime import datetime, timezone
from difflib import SequenceMatcher
from typing import Any, Iterable
from uuid import uuid4
from xml.etree import ElementTree

from botocore.exceptions import BotoCoreError, ClientError
from flask import current_app
from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage

from meeting_assistant.services.knowledge_service import KnowledgeService
from meeting_assistant.utils.exceptions import DatabaseError, ResourceNotFoundError, ValidationError

_QUESTION_HEADERS = {"question", "questions", "interview question", "prompt"}
_ANSWER_HEADERS = {"answer", "answers", "prepared answer", "response", "suggested answer"}

# Application Materials are indexed when saved so Career Evidence Search and
# Mock Interview do not repeatedly download and parse selected documents.
_MATERIAL_INDEX_VERSION = 2
_MAX_MATERIAL_INDEX_ENTRIES = 180
_MAX_MATERIAL_INDEX_BYTES = 180_000
_MAX_MATERIAL_ENTRY_CHARACTERS = 3_000
_MAX_FACT_CHUNK_CHARACTERS = 1_400

_CONTEXT_STOP_WORDS = {
    "a", "about", "an", "and", "are", "as", "at", "be", "been", "being",
    "by", "can", "could", "did", "do", "does", "for", "from", "had", "has",
    "have", "how", "i", "in", "is", "it", "made", "make", "me", "of", "on",
    "or", "our", "please", "regarding", "that", "the", "this", "to", "we",
    "were", "what", "when", "where", "which", "who", "why", "with", "would",
    "you", "your",
}

_CONTEXT_TOKEN_ALIASES = {
    "revenues": "revenue",
    "sales": "sale",
    "sold": "sale",
    "selling": "sale",
    "costs": "cost",
    "expenses": "expense",
    "units": "unit",
    "prices": "price",
    "profits": "profit",
    "margins": "margin",
}

_FINANCIAL_QUERY_TOKENS = {"revenue", "sale", "cost", "expense", "price", "profit", "margin"}
_FINANCIAL_CONTEXT_TOKENS = _FINANCIAL_QUERY_TOKENS | {"unit", "quantity", "amount", "total"}

# Question matching intentionally ignores low-information wording and normalizes
# common interview paraphrases. This lets a prepared question such as
# "What is your expected salary?" match "What salary do you expect?" without
# treating every question containing the word "salary" as equivalent.
_QUESTION_STOP_WORDS = {
    "a", "about", "an", "and", "are", "as", "at", "be", "being", "can", "could",
    "biggest", "describe", "did", "do", "does", "for", "from", "greatest",
    "have", "how", "i", "in",
    "is", "it", "main", "me", "most", "my", "of", "on", "or", "please",
    "say", "strongest", "tell", "that",
    "the", "this", "through", "to", "us", "walk", "was", "were", "what",
    "when", "where",
    "which", "who", "why", "should", "will", "willing", "with", "would",
    "you", "your",
}

_TOKEN_ALIASES = {
    "compensation": "salary",
    "pay": "salary",
    "remuneration": "salary",
    "expected": "expect",
    "expectation": "expect",
    "expectations": "expect",
    "expecting": "expect",
    "desired": "expect",
    "targeted": "target",
    "targeting": "target",
    "seeking": "seek",
    "looking": "look",
    "currently": "current",
    "now": "current",
    "presently": "current",
    "previous": "previous",
    "previously": "previous",
    "prior": "previous",
    "past": "previous",
    "negotiable": "negotiate",
    "negotiation": "negotiate",
    "negotiating": "negotiate",
    "flexibility": "flexible",
    "strengths": "strength",
    "quality": "strength",
    "qualities": "strength",
    "weaknesses": "weakness",
    "positions": "role",
    "position": "role",
    "jobs": "role",
    "job": "role",
    "companies": "company",
    "organizations": "company",
    "organisation": "company",
    "organisations": "company",
    "employer": "company",
    "employers": "company",
    "skills": "skill",
    "qualifications": "qualification",
    "requirements": "requirement",
    "responsibilities": "responsibility",
    "motivated": "motivate",
    "motivation": "motivate",
    "interested": "interest",
    "interests": "interest",
    "yourself": "self_intro",
    "background": "self_intro",
}

# Each family contains mutually different intents. A candidate is rejected when
# the asked question and prepared question clearly select different members of
# the same family. This is especially important for salary questions.
_INTENT_FAMILIES = {
    "salary_stage": {
        "expected": {"expect", "target", "seek", "look", "want"},
        "current": {"current"},
        "previous": {"previous", "former", "last"},
        "negotiation": {"negotiate", "flexible"},
    },
    "personal_trait": {
        "strength": {"strength"},
        "weakness": {"weakness"},
    },
}


class ApplicationMaterialsService:
    """Application-owned files, context, and prepared-answer index.

    ``ApplicationRecord`` is the aggregate root.  This service never creates a
    second meeting/workspace entity; all mutable materials are persisted as the
    ``APPLICATION_MATERIALS#<application_id>`` linked record in the canonical
    Career Bridge applications table.
    """

    def __init__(self, *, application_store: Any | None = None) -> None:
        self.knowledge = KnowledgeService()
        self.repository = self.knowledge.repository
        self.file_store = self.knowledge.file_store
        self.applications = application_store or current_app.extensions[
            "career_bridge_application_store"
        ]

    def list_applications(
        self, user_id: str, *, include_archived: bool = False
    ) -> list[dict[str, Any]]:
        try:
            applications = self.applications.list_for_owner(user_id)
            active_id = self.get_active_application_id(user_id)
        except (BotoCoreError, ClientError, OSError, ValueError) as exc:
            raise DatabaseError("Job applications could not be loaded.") from exc
        results: list[dict[str, Any]] = []
        for application in applications:
            if not include_archived and str(getattr(application, "status", "")) == "archived":
                continue
            payload = self._payload_for_application(user_id, application)
            results.append(self._serialize_application(application, payload, active_id))
        results.sort(
            key=lambda item: (
                item.get("scheduled_at") or "9999-12-31",
                item.get("updated_at") or "",
            )
        )
        return results

    def create_application(self, user_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        title = self.knowledge._required_text(
            payload.get("role") or payload.get("title"), "Job title", 240
        )
        parsed_role, parsed_company = _split_application_title(title)
        company = str(
            payload.get("company") or parsed_company or "Company not specified"
        ).strip()[:240]
        scheduled_at = str(payload.get("scheduled_at") or "").strip()
        application = self.applications.create(
            user_id,
            company=company or "Company not specified",
            role=parsed_role or title,
            interview_audience=", ".join(
                _string_list(payload.get("participants"), 100, 200)
            ),
            status="interviewing" if scheduled_at else "draft",
            notes=str(payload.get("purpose") or "").strip()[:2000],
            upcoming_event_date=_date_only(scheduled_at),
            upcoming_event_type="interview" if scheduled_at else "",
            job_description=str(payload.get("job_description") or "").strip(),
            workflow_step="setup",
        )
        if bool(payload.get("activate", True)):
            self.set_active_application(user_id, application.id)
        material_payload = self._default_payload(application)
        material_payload["interview_scheduled_at"] = scheduled_at
        material_payload["recruiter_contacts"] = _string_list(
            payload.get("participants"), 100, 200
        )
        self._save(material_payload)
        return self._serialize_application(
            application,
            material_payload,
            application.id if payload.get("activate", True) else self.get_active_application_id(user_id),
        )

    def update_application(
        self, user_id: str, application_id: str, payload: dict[str, Any]
    ) -> dict[str, Any]:
        application = self._application(user_id, application_id)
        application_id = str(getattr(application, "id", "") or "")
        material_payload = self._get(user_id, application_id)
        scheduled_at = str(payload.get("scheduled_at") or "").strip()
        participants = _string_list(payload.get("participants"), 100, 200)
        role = str(getattr(application, "role", "") or "")
        company = str(getattr(application, "company", "") or "")
        if "title" in payload:
            parsed_role, parsed_company = _split_application_title(
                str(payload.get("title") or "")
            )
            role = parsed_role or role
            company = parsed_company or company
        updated = self.applications.update(
            user_id,
            application_id,
            company=company,
            role=role,
            job_url=str(getattr(application, "job_url", "") or ""),
            application_date=str(getattr(application, "application_date", "") or ""),
            status=str(getattr(application, "status", "") or "draft"),
            screening_received=bool(getattr(application, "screening_received", False)),
            interview_received=bool(getattr(application, "interview_received", False)),
            offer_received=bool(getattr(application, "offer_received", False)),
            notes=(
                str(payload.get("purpose") or "").strip()[:2000]
                if "purpose" in payload
                else str(getattr(application, "notes", "") or "")
            ),
            next_follow_up_date=str(
                getattr(application, "next_follow_up_date", "") or ""
            ),
            interview_readiness=getattr(application, "interview_readiness", None),
            next_action=str(getattr(application, "next_action", "") or ""),
            upcoming_event_date=(
                _date_only(scheduled_at)
                if "scheduled_at" in payload
                else str(getattr(application, "upcoming_event_date", "") or "")
            ),
            upcoming_event_type=(
                "interview" if scheduled_at else ""
                if "scheduled_at" in payload
                else str(getattr(application, "upcoming_event_type", "") or "")
            ),
            job_description=str(getattr(application, "job_description", "") or ""),
            interview_audience=(
                ", ".join(participants)
                if "participants" in payload
                else str(getattr(application, "interview_audience", "") or "")
            ),
        )
        if updated is None:
            raise ResourceNotFoundError("Job application not found.")
        if "scheduled_at" in payload:
            material_payload["interview_scheduled_at"] = scheduled_at
        if "participants" in payload:
            material_payload["recruiter_contacts"] = participants
        material_payload["updated_at"] = _utc_now()
        self._save(material_payload)
        if payload.get("activate") is True:
            self.set_active_application(user_id, application_id)
        material_payload = self._get(user_id, application_id)
        return self._serialize_application(
            updated, material_payload, self.get_active_application_id(user_id)
        )

    def delete_application(self, user_id: str, application_id: str) -> dict[str, Any]:
        application = self._application(user_id, application_id)
        application_id = str(getattr(application, "id", "") or "")
        payload = self._get(user_id, application_id)
        for record in payload.get("temporary_files") or []:
            object_key = str(record.get("object_key") or "").strip()
            if not object_key:
                continue
            try:
                self.file_store.delete(object_key)
            except Exception:
                current_app.logger.exception(
                    "Could not delete application-only file for application %s",
                    application_id,
                )
        if not self.applications.delete(user_id, application_id):
            raise ResourceNotFoundError("Job application not found.")
        return self._serialize_application(application, payload)

    def delete_materials(self, user_id: str, application_id: str) -> None:
        """Remove application-only file bytes before the aggregate is deleted."""
        try:
            payload = self._get(user_id, application_id)
        except ResourceNotFoundError:
            return
        for record in payload.get("temporary_files") or []:
            object_key = str(record.get("object_key") or "").strip()
            if not object_key:
                continue
            try:
                self.file_store.delete(object_key)
            except Exception:
                current_app.logger.exception(
                    "Could not delete application-only file for application %s",
                    application_id,
                )

    def get_materials(self, user_id: str, application_id: str) -> dict[str, Any]:
        return self._serialize_materials(self._get(user_id, application_id))

    def save_materials(self, user_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        application_id = _payload_application_id(payload)
        item = self._get(user_id, application_id)
        requested_ids = list(
            dict.fromkeys(
                _string_list(
                    payload.get("library_file_ids") or payload.get("file_ids"),
                    500,
                    160,
                )
            )
        )
        valid_ids = []
        for file_id in requested_ids:
            if self.repository.get_file(user_id, file_id):
                valid_ids.append(file_id)
            else:
                raise ValidationError("One of the selected documents no longer exists.")
        item["library_file_ids"] = valid_ids
        context = payload.get("application_context")
        if context is None:
            context = payload.get("meeting_context")
        if isinstance(context, dict):
            item["application_context"] = _clean_context(context)
        if "recruiter_contacts" in payload or "participants" in payload:
            item["recruiter_contacts"] = _string_list(
                payload.get("recruiter_contacts") or payload.get("participants"),
                100,
                200,
            )
        if "recruiter_messages" in payload or "messages" in payload:
            item["recruiter_messages"] = _clean_recruiter_messages(
                payload.get("recruiter_messages") or payload.get("messages")
            )
        if "interview_scheduled_at" in payload or "scheduled_at" in payload:
            item["interview_scheduled_at"] = str(
                payload.get("interview_scheduled_at")
                or payload.get("scheduled_at")
                or ""
            ).strip()[:120]
        self._rebuild_material_index(user_id, item)
        item["updated_at"] = _utc_now()
        self._save(item)
        if payload.get("activate", True):
            self.set_active_application(user_id, application_id)
        return self._serialize_materials(item)

    def save_application_context(
        self, user_id: str, application_id: str, context: dict[str, Any]
    ) -> dict[str, Any]:
        item = self._get(user_id, application_id)
        item["application_context"] = _clean_context(context)
        item["updated_at"] = _utc_now()
        self._save(item)
        return dict(item["application_context"])

    def upload_temporary_files(
        self,
        user_id: str,
        application_id: str,
        uploads: Iterable[FileStorage],
    ) -> list[dict[str, Any]]:
        item = self._get(user_id, application_id)
        selected = [upload for upload in uploads if upload and upload.filename]
        if not selected:
            raise ValidationError("Choose at least one application-only file.")
        temporary = list(item.get("temporary_files") or [])
        added: list[dict[str, Any]] = []
        try:
            for upload in selected:
                original_name, stored_name, extension, content_type, content = (
                    self.knowledge._prepare_upload(upload)
                )
                file_id = uuid4().hex
                object_key = self._temporary_object_key(
                    user_id,
                    str(item.get("application_id") or application_id),
                    file_id,
                    stored_name,
                )
                self.file_store.put(object_key, content, content_type)
                record = {
                    "file_id": file_id,
                    "id": file_id,
                    "name": original_name,
                    "filename": original_name,
                    "extension": extension,
                    "content_type": content_type,
                    "size_bytes": len(content),
                    "object_key": object_key,
                    "added_at": _utc_now(),
                }
                temporary.append(record)
                added.append(record)
            item["temporary_files"] = temporary
            self._rebuild_material_index(user_id, item)
            item["updated_at"] = _utc_now()
            self._save(item)
        except Exception:
            for record in added:
                try:
                    self.file_store.delete(record["object_key"])
                except Exception:
                    current_app.logger.exception(
                        "Could not roll back application-only file"
                    )
            raise
        return [self._serialize_temporary_file(record) for record in added]

    def delete_temporary_file(
        self, user_id: str, application_id: str, file_id: str
    ) -> None:
        item = self._get(user_id, application_id)
        files = list(item.get("temporary_files") or [])
        target = next(
            (
                record
                for record in files
                if str(record.get("file_id") or record.get("id")) == file_id
            ),
            None,
        )
        if not target:
            raise ResourceNotFoundError("Application-only file not found.")
        item["temporary_files"] = [record for record in files if record is not target]
        self._rebuild_material_index(user_id, item)
        item["updated_at"] = _utc_now()
        self._save(item)
        self.file_store.delete(str(target.get("object_key") or ""))

    def clear_temporary_files(self, user_id: str, application_id: str) -> None:
        item = self._get(user_id, application_id)
        files = list(item.get("temporary_files") or [])
        item["temporary_files"] = []
        self._rebuild_material_index(user_id, item)
        item["updated_at"] = _utc_now()
        self._save(item)
        for record in files:
            try:
                self.file_store.delete(str(record.get("object_key") or ""))
            except Exception:
                current_app.logger.exception("Could not delete application-only file")

    def get_active_application_id(self, user_id: str) -> str:
        try:
            active_id = self.applications.get_active_application_id(user_id)
        except (BotoCoreError, ClientError, OSError, ValueError) as exc:
            raise DatabaseError("The active application could not be loaded.") from exc
        return active_id or ""

    def set_active_application(self, user_id: str, application_id: str) -> str:
        normalized = _normalize_application_id(application_id)
        try:
            return self.applications.set_active_application_id(user_id, normalized)
        except ValueError as exc:
            raise ValidationError(str(exc)) from exc
        except (BotoCoreError, ClientError, OSError) as exc:
            raise DatabaseError("The active application could not be saved.") from exc

    def complete_interview(
        self, user_id: str, application_id: str, completed_interview_id: str
    ) -> None:
        application_id = _normalize_application_id(application_id)
        if not application_id:
            return
        item = self._get(user_id, application_id)
        now = _utc_now()
        item.update(
            {
                "last_completed_interview_at": now,
                "last_completed_interview_id": completed_interview_id,
                "updated_at": now,
            }
        )
        self._save(item)
        if self.get_active_application_id(user_id) == application_id:
            self.set_active_application(user_id, "")

    def find_prepared_answer(
        self, user_id: str, question: str, application_id: str = ""
    ) -> dict[str, Any] | None:
        selected_id, item = self._resolve_application(user_id, application_id)
        if not item:
            return None
        normalized_question = _normalize_question(question)
        if not normalized_question:
            return None
        candidates = [
            entry
            for entry in self._ensure_material_index(user_id, item)
            if str(entry.get("source_type") or "") == "prepared_answer"
        ]
        ranked: list[tuple[float, str, dict[str, Any]]] = []
        for candidate in candidates:
            score, confidence = _question_match(
                normalized_question, str(candidate.get("question") or "")
            )
            if confidence != "none":
                ranked.append((score, confidence, candidate))
        if not ranked:
            return None
        ranked.sort(key=lambda value: value[0], reverse=True)
        score, confidence, candidate = ranked[0]
        if confidence == "high" and len(ranked) > 1:
            second_score, _, second_candidate = ranked[1]
            different_question = _normalize_question(
                candidate.get("question", "")
            ) != _normalize_question(second_candidate.get("question", ""))
            if different_question and score - second_score < 0.04:
                confidence = "medium"
        title = str(item.get("application_title") or "")
        return {
            **candidate,
            "application_id": selected_id,
            "meeting_id": selected_id,
            "application_title": title,
            "meeting_title": title,
            "match_score": round(score, 4),
            "match_confidence": confidence,
            "use_verbatim": confidence == "high",
        }

    def find_relevant_context(
        self,
        user_id: str,
        question: str,
        application_id: str = "",
        *,
        limit: int = 4,
    ) -> list[dict[str, Any]]:
        selected_id, item = self._resolve_application(user_id, application_id)
        if not item or not str(question or "").strip():
            return []
        ranked: list[tuple[float, dict[str, Any]]] = []
        for entry in self._ensure_material_index(user_id, item):
            if str(entry.get("source_type") or "") not in {
                "document_context",
                "prepared_answer",
            }:
                continue
            score = _context_match_score(question, entry)
            if score > 0:
                ranked.append((score, entry))
        ranked.sort(key=lambda value: value[0], reverse=True)
        title = str(item.get("application_title") or "")
        results: list[dict[str, Any]] = []
        seen_text: set[str] = set()
        for score, entry in ranked:
            context_text = _entry_context_text(entry)
            normalized_text = _normalize_question(context_text)
            if not context_text or normalized_text in seen_text:
                continue
            seen_text.add(normalized_text)
            results.append(
                {
                    **entry,
                    "text": context_text,
                    "application_id": selected_id,
                    "meeting_id": selected_id,
                    "application_title": title,
                    "meeting_title": title,
                    "match_score": round(score, 4),
                }
            )
            if len(results) >= max(1, min(int(limit), 8)):
                break
        return results

    def _resolve_application(
        self, user_id: str, application_id: str = ""
    ) -> tuple[str, dict[str, Any] | None]:
        selected_id = (
            _normalize_application_id(application_id)
            or self.get_active_application_id(user_id)
        )
        if not selected_id:
            return "", None
        try:
            return selected_id, self._get(user_id, selected_id)
        except ResourceNotFoundError:
            return selected_id, None

    def _ensure_material_index(
        self, user_id: str, item: dict[str, Any]
    ) -> list[dict[str, Any]]:
        if (
            int(item.get("material_index_version") or 0)
            == _MATERIAL_INDEX_VERSION
            and isinstance(item.get("material_index"), list)
        ):
            return list(item.get("material_index") or [])
        self._rebuild_material_index(user_id, item)
        item["updated_at"] = _utc_now()
        self._save(item)
        return list(item.get("material_index") or [])

    def _rebuild_material_index(self, user_id: str, item: dict[str, Any]) -> None:
        entries: list[dict[str, Any]] = []
        total_bytes = 0

        def add_document(file_item: dict[str, Any], content: bytes) -> None:
            nonlocal total_bytes
            try:
                extracted = _extract_material_entries(file_item, content)
            except Exception:
                current_app.logger.exception(
                    "Could not index Application Material %s",
                    file_item.get("file_id")
                    or file_item.get("filename")
                    or file_item.get("name"),
                )
                return
            for raw_entry in extracted:
                entry = _sanitize_index_entry(raw_entry)
                entry_size = _index_entry_byte_count(entry)
                if not entry or entry_size <= 0:
                    continue
                if len(entries) >= _MAX_MATERIAL_INDEX_ENTRIES:
                    return
                if total_bytes + entry_size > _MAX_MATERIAL_INDEX_BYTES:
                    return
                entries.append(entry)
                total_bytes += entry_size

        for file_id in item.get("library_file_ids") or []:
            file_item = self.repository.get_file(user_id, str(file_id))
            if not file_item:
                continue
            try:
                content = self.file_store.get(str(file_item.get("object_key") or ""))
            except Exception:
                current_app.logger.exception(
                    "Could not load Application Material %s", file_id
                )
                continue
            add_document(file_item, content)
        for file_item in item.get("temporary_files") or []:
            try:
                content = self.file_store.get(str(file_item.get("object_key") or ""))
            except Exception:
                current_app.logger.exception(
                    "Could not load application-only material"
                )
                continue
            add_document(file_item, content)
        item["material_index"] = entries
        item["material_index_version"] = _MATERIAL_INDEX_VERSION
        item["material_indexed_at"] = _utc_now()

    def _application(self, user_id: str, application_id: str) -> Any:
        normalized = _normalize_application_id(application_id)
        if not normalized:
            raise ValidationError("application_id is required.")
        try:
            application = self.applications.get(
                user_id, normalized, include_resume_bytes=False
            )
        except (BotoCoreError, ClientError, OSError) as exc:
            raise DatabaseError("The job application could not be loaded.") from exc
        if application is None:
            raise ResourceNotFoundError("Job application not found.")
        return application

    def _get(self, user_id: str, application_id: str) -> dict[str, Any]:
        application = self._application(user_id, application_id)
        return self._payload_for_application(user_id, application)

    def _payload_for_application(
        self, user_id: str, application: Any
    ) -> dict[str, Any]:
        application_id = str(getattr(application, "id", "") or "")
        try:
            record = self.applications.get_application_materials(
                user_id, application_id
            )
        except (BotoCoreError, ClientError, OSError, ValueError) as exc:
            raise DatabaseError("Application Materials could not be loaded.") from exc
        if record is None:
            return self._default_payload(application)
        payload = dict(record.payload())
        payload.setdefault("created_at", record.created_at)
        payload.setdefault("updated_at", record.updated_at)
        return self._normalize_payload(application, payload)

    def _save(self, item: dict[str, Any]) -> None:
        application_id = str(item.get("application_id") or "").strip()
        if not application_id:
            raise ValidationError("application_id is required.")
        persisted = {
            key: value
            for key, value in item.items()
            if key not in {"application_title", "company", "role", "status"}
        }
        try:
            self.applications.save_application_materials(
                str(item.get("owner_id") or ""),
                application_id,
                payload_json=json.dumps(
                    persisted,
                    ensure_ascii=False,
                    separators=(",", ":"),
                    default=str,
                ),
            )
        except ValueError as exc:
            raise ValidationError(str(exc)) from exc
        except (BotoCoreError, ClientError, OSError) as exc:
            raise DatabaseError("Application Materials could not be saved.") from exc

    def _default_payload(self, application: Any) -> dict[str, Any]:
        now = _utc_now()
        return self._normalize_payload(
            application,
            {
                "application_id": str(getattr(application, "id", "") or ""),
                "owner_id": str(getattr(application, "owner_id", "") or ""),
                "library_file_ids": [],
                "temporary_files": [],
                "application_context": {},
                "interview_scheduled_at": (
                    str(getattr(application, "upcoming_event_date", "") or "")
                    if str(getattr(application, "upcoming_event_type", "") or "") == "interview"
                    else ""
                ),
                "recruiter_contacts": _participant_values(
                    str(getattr(application, "interview_audience", "") or "")
                ),
                "recruiter_messages": [],
                "material_index": [],
                "material_index_version": _MATERIAL_INDEX_VERSION,
                "material_indexed_at": now,
                "created_at": str(getattr(application, "created_at", "") or now),
                "updated_at": str(getattr(application, "updated_at", "") or now),
                "last_completed_interview_at": "",
                "last_completed_interview_id": "",
            },
        )

    def _normalize_payload(
        self, application: Any, payload: dict[str, Any]
    ) -> dict[str, Any]:
        application_id = str(getattr(application, "id", "") or "")
        normalized = dict(payload or {})
        normalized["application_id"] = application_id
        normalized["owner_id"] = str(getattr(application, "owner_id", "") or "")
        normalized["application_title"] = _application_title(application)
        normalized["company"] = str(getattr(application, "company", "") or "")
        normalized["role"] = str(getattr(application, "role", "") or "")
        normalized["status"] = str(getattr(application, "status", "") or "draft")
        normalized["library_file_ids"] = list(
            normalized.get("library_file_ids") or []
        )
        normalized["temporary_files"] = list(
            normalized.get("temporary_files") or []
        )
        context = normalized.get("application_context")
        if context is None:
            context = normalized.get("meeting_context")
        normalized["application_context"] = _clean_context(context or {})
        normalized.setdefault("interview_scheduled_at", "")
        normalized["recruiter_contacts"] = list(
            normalized.get("recruiter_contacts") or []
        )
        normalized["recruiter_messages"] = list(
            normalized.get("recruiter_messages") or []
        )
        normalized.setdefault("material_index", [])
        normalized.setdefault("material_index_version", 0)
        normalized.setdefault("material_indexed_at", "")
        normalized.setdefault("created_at", str(getattr(application, "created_at", "") or ""))
        normalized.setdefault("updated_at", str(getattr(application, "updated_at", "") or ""))
        return normalized

    @staticmethod
    def _serialize_application(
        application: Any, payload: dict[str, Any], active_id: str = ""
    ) -> dict[str, Any]:
        application_id = str(getattr(application, "id", "") or "")
        interview_audience = str(
            getattr(application, "interview_audience", "") or ""
        )
        return {
            "id": application_id,
            "application_id": application_id,
            "meeting_id": application_id,
            "source": "application_builder",
            "title": _application_title(application),
            "company": str(getattr(application, "company", "") or ""),
            "role": str(getattr(application, "role", "") or ""),
            "scheduled_at": str(
                payload.get("interview_scheduled_at")
                or (
                    str(getattr(application, "upcoming_event_date", "") or "")
                    if str(getattr(application, "upcoming_event_type", "") or "") == "interview"
                    else ""
                )
            ),
            "participants": list(
                payload.get("recruiter_contacts")
                or _participant_values(interview_audience)
            ),
            "purpose": str(getattr(application, "notes", "") or ""),
            "status": str(getattr(application, "status", "") or "draft"),
            "created_at": str(getattr(application, "created_at", "") or ""),
            "updated_at": str(getattr(application, "updated_at", "") or ""),
            "completed_at": str(
                payload.get("last_completed_interview_at") or ""
            ),
            "completed_meeting_id": str(
                payload.get("last_completed_interview_id") or ""
            ),
            "active": application_id == active_id,
            "library_file_count": len(payload.get("library_file_ids") or []),
            "temporary_file_count": len(payload.get("temporary_files") or []),
        }

    def _serialize_materials(self, item: dict[str, Any]) -> dict[str, Any]:
        application_id = str(item.get("application_id") or "")
        context = _clean_context(item.get("application_context") or {})
        return {
            "application_id": application_id,
            "application_workspace_id": application_id,
            "meeting_id": application_id,
            "library_file_ids": list(item.get("library_file_ids") or []),
            "temporary_files": [
                self._serialize_temporary_file(record)
                for record in item.get("temporary_files") or []
            ],
            "application_context": context,
            "meeting_context": context,
            "interview_scheduled_at": str(
                item.get("interview_scheduled_at") or ""
            ),
            "recruiter_contacts": list(item.get("recruiter_contacts") or []),
            "recruiter_messages": list(item.get("recruiter_messages") or []),
        }

    @staticmethod
    def _serialize_temporary_file(record: dict[str, Any]) -> dict[str, Any]:
        file_id = str(record.get("file_id") or record.get("id") or "")
        return {
            "id": file_id,
            "file_id": file_id,
            "name": str(
                record.get("name") or record.get("filename") or "Application-only file"
            ),
            "filename": str(
                record.get("filename") or record.get("name") or "Application-only file"
            ),
            "extension": str(record.get("extension") or ""),
            "content_type": str(record.get("content_type") or ""),
            "size_bytes": int(record.get("size_bytes") or 0),
            "added_at": str(record.get("added_at") or ""),
        }

    @staticmethod
    def _temporary_object_key(
        user_id: str,
        application_id: str,
        file_id: str,
        stored_name: str,
    ) -> str:
        import hashlib

        owner_hash = hashlib.sha256(user_id.encode("utf-8")).hexdigest()[:24]
        return (
            f"knowledge/{owner_hash}/applications/{application_id}/"
            f"temporary/{file_id}/{stored_name}"
        )


def _clean_recruiter_messages(value: Any) -> list[dict[str, str]]:
    """Normalize recruiter communication records stored under an application."""

    raw_items = value if isinstance(value, list) else ([value] if value else [])
    messages: list[dict[str, str]] = []
    for raw in raw_items[:200]:
        if isinstance(raw, dict):
            body = str(
                raw.get("body") or raw.get("message") or raw.get("text") or ""
            ).strip()[:10_000]
            if not body:
                continue
            messages.append(
                {
                    "id": str(raw.get("id") or raw.get("message_id") or "").strip()[:160],
                    "subject": str(raw.get("subject") or "").strip()[:500],
                    "body": body,
                    "sender": str(raw.get("sender") or raw.get("from") or "").strip()[:320],
                    "recipient": str(raw.get("recipient") or raw.get("to") or "").strip()[:320],
                    "direction": str(raw.get("direction") or "").strip()[:40],
                    "channel": str(raw.get("channel") or "").strip()[:80],
                    "sent_at": str(raw.get("sent_at") or raw.get("created_at") or "").strip()[:120],
                }
            )
            continue
        body = str(raw or "").strip()[:10_000]
        if body:
            messages.append(
                {
                    "id": "",
                    "subject": "",
                    "body": body,
                    "sender": "",
                    "recipient": "",
                    "direction": "",
                    "channel": "",
                    "sent_at": "",
                }
            )
    return messages


def _normalize_application_id(value: Any) -> str:
    """Return the canonical application ID from supported UI aliases."""

    normalized = str(value or "").strip()
    prefix, separator, raw_id = normalized.partition(":")
    if separator and prefix in {"builder", "application"}:
        return raw_id.strip()
    return normalized


def _payload_application_id(payload: dict[str, Any]) -> str:
    application_id = _normalize_application_id(
        payload.get("application_id")
        or payload.get("application_workspace_id")
        or payload.get("meeting_id")
        or ""
    )
    if not application_id:
        raise ValidationError("application_id is required.")
    return application_id


def _application_title(application: Any) -> str:
    role = str(getattr(application, "role", "") or "").strip()
    company = str(getattr(application, "company", "") or "").strip()
    if role and company:
        return f"{role} — {company}"
    return role or company or "Untitled application"


def _split_application_title(value: str) -> tuple[str, str]:
    normalized = " ".join(str(value or "").split()).strip()
    if not normalized:
        return "", ""
    for separator in (" — ", " – ", " at "):
        if separator in normalized:
            role, company = normalized.split(separator, 1)
            return role.strip()[:240], company.strip()[:240]
    return normalized[:240], ""


def _participant_values(value: str) -> list[str]:
    return [
        item.strip()
        for item in re.split(r"[,;\n]+", str(value or ""))
        if item.strip()
    ][:100]


def _date_only(value: str) -> str:
    candidate = str(value or "").strip()
    if not candidate:
        return ""
    return candidate[:10] if re.match(r"^\d{4}-\d{2}-\d{2}", candidate) else ""


def _extract_material_entries(item: dict[str, Any], content: bytes) -> list[dict[str, Any]]:
    extension = str(item.get("extension") or "").lower()
    filename = str(item.get("display_name") or item.get("filename") or item.get("name") or "Document")
    prepared = _extract_candidates(item, content)
    if extension == "xlsx":
        facts = _xlsx_fact_chunks(filename, content)
    elif extension == "xls":
        facts = _xls_fact_chunks(filename, content)
    elif extension == "pdf":
        facts = _pdf_fact_chunks(filename, content)
    elif extension in {"txt", "md"}:
        facts = _text_fact_chunks(filename, content.decode("utf-8-sig", errors="replace"))
    elif extension == "docx":
        facts = _docx_fact_chunks(filename, content)
    else:
        facts = []
    return prepared + facts


def _extract_candidates(item: dict[str, Any], content: bytes) -> list[dict[str, Any]]:
    extension = str(item.get("extension") or "").lower()
    filename = str(item.get("display_name") or item.get("filename") or item.get("name") or "Document")
    if extension == "xlsx":
        return _xlsx_candidates(filename, content)
    if extension == "xls":
        return _xls_candidates(filename, content)
    if extension == "pdf":
        return _text_candidates(filename, _pdf_text(content))
    if extension in {"txt", "md"}:
        return _text_candidates(filename, content.decode("utf-8-sig", errors="replace"))
    if extension == "docx":
        return _text_candidates(filename, _docx_text(content))
    return []


def _xlsx_candidates(filename: str, content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    results: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            rows = [[str(value).strip() if value is not None else "" for value in row] for row in sheet.iter_rows(values_only=True)]
            question_index = answer_index = None
            start = 0
            for index, row in enumerate(rows[:20]):
                lowered = [value.casefold() for value in row]
                q = next((i for i, value in enumerate(lowered) if value in _QUESTION_HEADERS), None)
                a = next((i for i, value in enumerate(lowered) if value in _ANSWER_HEADERS), None)
                if q is not None and a is not None:
                    question_index, answer_index, start = q, a, index + 1
                    break
            for row_number, row in enumerate(rows[start:], start=start + 1):
                if question_index is not None and answer_index is not None:
                    question = row[question_index] if question_index < len(row) else ""
                    answer = row[answer_index] if answer_index < len(row) else ""
                else:
                    values = [value for value in row if value]
                    if len(values) < 2:
                        continue
                    question, answer = values[0], values[1]
                    if "?" not in question and not re.match(r"^(tell|describe|explain|why|how|what|when|where|who|give|walk)\b", question, re.I):
                        continue
                if question and answer:
                    results.append({"question": question, "answer": answer, "source_name": filename, "source_detail": f"{sheet.title} · row {row_number}", "source_type": "prepared_answer"})
    finally:
        workbook.close()
    return results



def _xls_candidates(filename: str, content: bytes) -> list[dict[str, Any]]:
    import xlrd
    workbook = xlrd.open_workbook(file_contents=content)
    results: list[dict[str, Any]] = []
    for sheet in workbook.sheets():
        rows = [[str(sheet.cell_value(row, column)).strip() for column in range(sheet.ncols)] for row in range(sheet.nrows)]
        question_index = answer_index = None
        start = 0
        for index, row in enumerate(rows[:20]):
            lowered = [value.casefold() for value in row]
            q = next((i for i, value in enumerate(lowered) if value in _QUESTION_HEADERS), None)
            a = next((i for i, value in enumerate(lowered) if value in _ANSWER_HEADERS), None)
            if q is not None and a is not None:
                question_index, answer_index, start = q, a, index + 1
                break
        for row_number, row in enumerate(rows[start:], start=start + 1):
            if question_index is not None and answer_index is not None:
                question = row[question_index] if question_index < len(row) else ""
                answer = row[answer_index] if answer_index < len(row) else ""
            else:
                values = [value for value in row if value]
                if len(values) < 2:
                    continue
                question, answer = values[0], values[1]
                if "?" not in question and not re.match(r"^(tell|describe|explain|why|how|what|when|where|who|give|walk)\b", question, re.I):
                    continue
            if question and answer:
                results.append({"question": question, "answer": answer, "source_name": filename, "source_detail": f"{sheet.name} · row {row_number}", "source_type": "prepared_answer"})
    return results


def _pdf_text(content: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _text_candidates(filename: str, text: str) -> list[dict[str, Any]]:
    results = []
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for index, line in enumerate(lines):
        match = re.match(r"^(?:q(?:uestion)?\s*[:\-]\s*)?(.+?\?)\s*(?:a(?:nswer)?\s*[:\-]\s*)?(.+)$", line, re.I)
        if match:
            results.append({"question": match.group(1).strip(), "answer": match.group(2).strip(), "source_name": filename, "source_detail": f"line {index + 1}", "source_type": "prepared_answer"})
        elif line.endswith("?") and index + 1 < len(lines):
            results.append({"question": line, "answer": lines[index + 1], "source_name": filename, "source_detail": f"lines {index + 1}-{index + 2}", "source_type": "prepared_answer"})
    return results


def _docx_text(content: bytes) -> str:
    import zipfile
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    return "\n".join(node.text or "" for node in root.iter() if node.tag.endswith("}t"))



def _xlsx_fact_chunks(filename: str, content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    results: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            headers: list[str] = []
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [_cell_text(value) for value in row]
                if not any(values):
                    continue
                if not headers and _looks_like_header_row(values):
                    headers = values
                text = _spreadsheet_row_text(values, headers if row_number > 1 else [])
                if text:
                    results.append(_fact_entry(filename, f"{sheet.title} · row {row_number}", text))
    finally:
        workbook.close()
    return results


def _xls_fact_chunks(filename: str, content: bytes) -> list[dict[str, Any]]:
    import xlrd
    workbook = xlrd.open_workbook(file_contents=content)
    results: list[dict[str, Any]] = []
    for sheet in workbook.sheets():
        headers: list[str] = []
        for row_index in range(sheet.nrows):
            values = [_cell_text(sheet.cell_value(row_index, column)) for column in range(sheet.ncols)]
            if not any(values):
                continue
            if not headers and _looks_like_header_row(values):
                headers = values
            text = _spreadsheet_row_text(values, headers if row_index > 0 else [])
            if text:
                results.append(_fact_entry(filename, f"{sheet.name} · row {row_index + 1}", text))
    return results


def _pdf_fact_chunks(filename: str, content: bytes) -> list[dict[str, Any]]:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(content))
    results: list[dict[str, Any]] = []
    for page_number, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""
        for chunk_number, chunk in enumerate(_chunk_text(page_text), start=1):
            detail = f"page {page_number}"
            if chunk_number > 1:
                detail += f" · part {chunk_number}"
            results.append(_fact_entry(filename, detail, chunk))
    return results


def _text_fact_chunks(filename: str, text: str) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    lines = text.splitlines()
    paragraph_lines: list[str] = []
    paragraph_start = 1

    def flush(end_line: int) -> None:
        nonlocal paragraph_lines, paragraph_start
        paragraph = " ".join(line.strip() for line in paragraph_lines if line.strip()).strip()
        if paragraph:
            for chunk_number, chunk in enumerate(_chunk_text(paragraph), start=1):
                detail = f"lines {paragraph_start}-{end_line}" if end_line != paragraph_start else f"line {paragraph_start}"
                if chunk_number > 1:
                    detail += f" · part {chunk_number}"
                results.append(_fact_entry(filename, detail, chunk))
        paragraph_lines = []

    for line_number, line in enumerate(lines, start=1):
        if line.strip():
            if not paragraph_lines:
                paragraph_start = line_number
            paragraph_lines.append(line)
        else:
            flush(line_number - 1)
    flush(len(lines))
    return results


def _docx_fact_chunks(filename: str, content: bytes) -> list[dict[str, Any]]:
    import zipfile
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    paragraphs: list[str] = []
    for paragraph in root.iter():
        if not paragraph.tag.endswith("}p"):
            continue
        text = "".join(node.text or "" for node in paragraph.iter() if node.tag.endswith("}t")).strip()
        if text:
            paragraphs.append(text)
    results: list[dict[str, Any]] = []
    for paragraph_number, paragraph in enumerate(paragraphs, start=1):
        for chunk_number, chunk in enumerate(_chunk_text(paragraph), start=1):
            detail = f"paragraph {paragraph_number}"
            if chunk_number > 1:
                detail += f" · part {chunk_number}"
            results.append(_fact_entry(filename, detail, chunk))
    return results


def _fact_entry(filename: str, detail: str, text: str) -> dict[str, Any]:
    return {
        "text": str(text or "").strip()[:_MAX_MATERIAL_ENTRY_CHARACTERS],
        "source_name": filename,
        "source_detail": detail,
        "source_type": "document_context",
    }


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_like_header_row(values: list[str]) -> bool:
    nonempty = [value for value in values if value]
    if not nonempty:
        return False
    alpha_cells = sum(bool(re.search(r"[A-Za-z]", value)) and not re.search(r"\d", value) for value in nonempty)
    return alpha_cells >= max(1, len(nonempty) // 2)


def _spreadsheet_row_text(values: list[str], headers: list[str]) -> str:
    pairs: list[str] = []
    for index, value in enumerate(values):
        if not value:
            continue
        header = headers[index].strip() if index < len(headers) else ""
        if header and header.casefold() != value.casefold():
            pairs.append(f"{header}: {value}")
        else:
            pairs.append(value)
    return "; ".join(pairs)[:_MAX_FACT_CHUNK_CHARACTERS]


def _chunk_text(text: str) -> list[str]:
    normalized = re.sub(r"\s+", " ", str(text or "")).strip()
    if not normalized:
        return []
    sentences = re.split(r"(?<=[.!?])\s+", normalized)
    chunks: list[str] = []
    current = ""
    for sentence in sentences:
        sentence = sentence.strip()
        if not sentence:
            continue
        if len(sentence) > _MAX_FACT_CHUNK_CHARACTERS:
            if current:
                chunks.append(current)
                current = ""
            for offset in range(0, len(sentence), _MAX_FACT_CHUNK_CHARACTERS):
                chunks.append(sentence[offset: offset + _MAX_FACT_CHUNK_CHARACTERS])
            continue
        candidate = f"{current} {sentence}".strip()
        if current and len(candidate) > _MAX_FACT_CHUNK_CHARACTERS:
            chunks.append(current)
            current = sentence
        else:
            current = candidate
    if current:
        chunks.append(current)
    return chunks


def _sanitize_index_entry(entry: dict[str, Any]) -> dict[str, Any]:
    source_type = str(entry.get("source_type") or "").strip()
    if source_type == "prepared_answer":
        question = str(entry.get("question") or "").strip()[:1_000]
        answer = str(entry.get("answer") or "").strip()[:_MAX_MATERIAL_ENTRY_CHARACTERS]
        if not question or not answer:
            return {}
        sanitized = {"question": question, "answer": answer, "source_type": source_type}
    elif source_type == "document_context":
        content = str(entry.get("text") or "").strip()[:_MAX_MATERIAL_ENTRY_CHARACTERS]
        if not content:
            return {}
        sanitized = {"text": content, "source_type": source_type}
    else:
        return {}
    sanitized["source_name"] = str(entry.get("source_name") or "Meeting material").strip()[:500]
    sanitized["source_detail"] = str(entry.get("source_detail") or "").strip()[:500]
    return sanitized


def _index_entry_byte_count(entry: dict[str, Any]) -> int:
    return sum(len(str(value).encode("utf-8")) for value in entry.values())


def _entry_context_text(entry: dict[str, Any]) -> str:
    if str(entry.get("source_type") or "") == "prepared_answer":
        question = str(entry.get("question") or "").strip()
        answer = str(entry.get("answer") or "").strip()
        return f"Question: {question}\nPrepared answer: {answer}".strip()
    return str(entry.get("text") or "").strip()


def _context_tokens(value: str) -> set[str]:
    tokens: set[str] = set()
    for token in re.findall(r"[A-Za-z0-9][A-Za-z0-9_.-]*", str(value or "").casefold()):
        if token in _CONTEXT_STOP_WORDS:
            continue
        canonical = _CONTEXT_TOKEN_ALIASES.get(token, token)
        if len(canonical) > 3 and canonical.endswith("s") and not canonical.endswith("ss"):
            canonical = canonical[:-1]
        tokens.add(_CONTEXT_TOKEN_ALIASES.get(canonical, canonical))
    return tokens


def _identifier_tokens(value: str) -> set[str]:
    return {
        token.casefold()
        for token in re.findall(r"\b(?=[A-Za-z0-9_-]*[A-Za-z])(?=[A-Za-z0-9_-]*\d)[A-Za-z0-9][A-Za-z0-9_-]{2,}\b", str(value or ""))
    }


def _context_match_score(question: str, entry: dict[str, Any]) -> float:
    context_text = _entry_context_text(entry)
    question_tokens = _context_tokens(question)
    context_tokens = _context_tokens(context_text)
    if not question_tokens or not context_tokens:
        return 0.0

    question_ids = _identifier_tokens(question)
    context_ids = _identifier_tokens(context_text)
    shared_ids = question_ids & context_ids
    if question_ids and not shared_ids:
        return 0.0

    overlap = question_tokens & context_tokens
    query_coverage = len(overlap) / len(question_tokens)
    context_precision = len(overlap) / min(len(context_tokens), 20)
    score = (0.56 * query_coverage) + (0.20 * context_precision)

    if shared_ids:
        # Exact product/project/account identifiers are stronger evidence than
        # broad vocabulary overlap. This is the key signal for BRAC204-like IDs.
        score += 0.55 + min(0.10, 0.05 * len(shared_ids))

    financial_query = bool(question_tokens & _FINANCIAL_QUERY_TOKENS)
    financial_context = bool(context_tokens & _FINANCIAL_CONTEXT_TOKENS)
    has_numeric_evidence = bool(re.search(r"(?:[$€£]\s*\d|\b\d+(?:[.,]\d+)?\b)", context_text))
    asks_amount = bool(re.search(r"\b(how much|total|amount|revenue|cost|price|profit|margin)\b", question, re.I))
    if financial_query and financial_context:
        score += 0.12
    if asks_amount and has_numeric_evidence:
        score += 0.12

    # Without a specific identifier, require meaningful lexical support to avoid
    # attaching generic but unrelated document passages.
    if not shared_ids and (query_coverage < 0.28 or len(overlap) < 2):
        return 0.0
    return min(score, 1.0)

def _normalize_question(value: str) -> str:
    text = re.sub(r"[^\w\s]", " ", str(value or "").casefold())
    return re.sub(r"\s+", " ", text).strip()


def _canonical_token(token: str) -> str:
    token = _TOKEN_ALIASES.get(token, token)
    if token in _TOKEN_ALIASES:
        return _TOKEN_ALIASES[token]

    # Conservative stemming for ordinary plurals and verb forms. Aliases are
    # checked again after stemming so words such as "expectations" normalize
    # predictably without requiring a heavyweight NLP dependency.
    stemmed = token
    if len(stemmed) > 5 and stemmed.endswith("ies"):
        stemmed = stemmed[:-3] + "y"
    elif len(stemmed) > 5 and stemmed.endswith("ing"):
        stemmed = stemmed[:-3]
    elif len(stemmed) > 4 and stemmed.endswith("ed"):
        stemmed = stemmed[:-2]
    elif len(stemmed) > 4 and stemmed.endswith("s") and not stemmed.endswith("ss"):
        stemmed = stemmed[:-1]
    return _TOKEN_ALIASES.get(stemmed, stemmed)


def _meaningful_question_tokens(value: str) -> list[str]:
    normalized = _normalize_question(value)
    tokens = [
        _canonical_token(token)
        for token in normalized.split()
        if token and token not in _QUESTION_STOP_WORDS
    ]

    # Salary expectation phrases have several natural forms. Canonicalize their
    # intent only when salary is actually present, preventing a phrase such as
    # "Why do you want this role?" from being interpreted as salary-related.
    if "salary" in tokens:
        salary_expectation_words = {"target", "seek", "look", "want"}
        if any(token in salary_expectation_words for token in tokens):
            tokens = [
                "expect" if token in salary_expectation_words else token
                for token in tokens
            ]
        if "flexible" in tokens:
            tokens = ["negotiate" if token == "flexible" else token for token in tokens]
        if not any(token in {"current", "previous", "negotiate"} for token in tokens):
            if any(token in {"range", "requirement"} for token in tokens):
                tokens.append("expect")

    if "role" in tokens and any(token in {"want", "motivate", "interest"} for token in tokens):
        tokens = [
            "interest" if token in {"want", "motivate", "interest"} else token
            for token in tokens
        ]

    # Preserve order for the secondary sequence signal while removing duplicates.
    return list(dict.fromkeys(tokens))


def _question_intents(tokens: set[str]) -> dict[str, set[str]]:
    intents: dict[str, set[str]] = {}
    for family, labels in _INTENT_FAMILIES.items():
        if family == "salary_stage" and "salary" not in tokens:
            continue
        matched = {
            label
            for label, vocabulary in labels.items()
            if tokens.intersection(vocabulary)
        }
        if matched:
            intents[family] = matched
    return intents


def _intent_relationship(question_tokens: set[str], candidate_tokens: set[str]) -> tuple[bool, bool]:
    question_intents = _question_intents(question_tokens)
    candidate_intents = _question_intents(candidate_tokens)
    partial = False
    for family in question_intents.keys() & candidate_intents.keys():
        question_labels = question_intents[family]
        candidate_labels = candidate_intents[family]
        if question_labels.isdisjoint(candidate_labels):
            return True, False
        if question_labels != candidate_labels:
            partial = True
    return False, partial


def _question_match(question: str, candidate_question: str) -> tuple[float, str]:
    normalized_question = _normalize_question(question)
    normalized_candidate = _normalize_question(candidate_question)
    if not normalized_question or not normalized_candidate:
        return 0.0, "none"
    if normalized_question == normalized_candidate:
        return 1.0, "high"

    question_tokens = _meaningful_question_tokens(normalized_question)
    candidate_tokens = _meaningful_question_tokens(normalized_candidate)
    question_set = set(question_tokens)
    candidate_set = set(candidate_tokens)

    if not question_set or not candidate_set:
        score = SequenceMatcher(None, normalized_question, normalized_candidate).ratio()
        return (score, "medium") if score >= 0.78 else (score, "none")

    intent_conflict, partial_intent = _intent_relationship(question_set, candidate_set)
    if intent_conflict:
        return 0.0, "none"

    intersection = question_set & candidate_set
    if not intersection:
        return 0.0, "none"

    containment = len(intersection) / min(len(question_set), len(candidate_set))
    jaccard = len(intersection) / len(question_set | candidate_set)
    character_score = SequenceMatcher(None, normalized_question, normalized_candidate).ratio()
    token_order_score = SequenceMatcher(None, question_tokens, candidate_tokens).ratio()
    score = (0.50 * containment) + (0.25 * jaccard) + (0.15 * character_score) + (0.10 * token_order_score)

    # Identical semantic token sets are high confidence even when the words are
    # reordered. This is the important case for "expected salary" versus
    # "salary do you expect". Two or more strongly overlapping core tokens also
    # qualify when no conflicting intent was detected.
    if not partial_intent and question_set == candidate_set:
        return max(score, 0.90), "high"
    if not partial_intent and len(intersection) >= 2 and containment >= 0.80 and jaccard >= 0.55:
        return score, "high"
    if not partial_intent and score >= 0.84 and containment >= 0.70:
        return score, "high"

    # Medium matches remain useful, but should be passed to the model as an
    # authoritative prepared answer rather than returned blindly.
    if score >= 0.58 and containment >= 0.50:
        return score, "medium"
    return score, "none"


def _string_list(value: Any, maximum_items: int, maximum_length: int) -> list[str]:
    values = value if isinstance(value, list) else str(value or "").split(",")
    return [str(item).strip()[:maximum_length] for item in values if str(item).strip()][:maximum_items]


def _clean_context(value: dict[str, Any]) -> dict[str, Any]:
    raw = value if isinstance(value, dict) else {}
    special_instructions = str(
        raw.get("special_instructions")
        or raw.get("meeting_instructions")
        or raw.get("instructions")
        or ""
    ).strip()

    # Preserve meeting contexts saved by older clients that used three
    # separate narrative fields. They are combined into the new single field
    # the first time the context is read or saved.
    if not special_instructions:
        legacy_parts: list[str] = []
        topics = str(raw.get("topics") or "").strip()
        constraints = str(raw.get("constraints") or "").strip()
        free_text = str(raw.get("free_text") or "").strip()
        if topics:
            legacy_parts.append(f"Topics to prioritize: {topics}")
        if constraints:
            legacy_parts.append(f"Constraints or sensitivities: {constraints}")
        if free_text:
            legacy_parts.append(free_text)
        special_instructions = "\n".join(legacy_parts)

    return {
        "objective": str(raw.get("objective") or "").strip()[:4000],
        "participants": str(raw.get("participants") or "").strip()[:4000],
        "special_instructions": special_instructions[:4000],
    }


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()
